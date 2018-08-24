# http://openpyxl.readthedocs.io/en/stable/
import openpyxl
from city_table import city_post_table, PathTable


log = print


def add_province():
    outbook = openpyxl.Workbook()
    outsheet = outbook.active
    outsheet.title = '加入省级区域'

    filename = PathTable.excelin
    inbook = openpyxl.load_workbook(filename)
    insheet = inbook['city']
    max_column = insheet.max_column
    max_row = insheet.max_row
    citys = city_post_table()[0]

    for i, inrows in enumerate(insheet.rows):
        outc = outsheet.cell(row=i+1, column=3)
        for j in range(5):
            if j < 2:
                inc = inrows[j]
                outc.value = inc.value
            elif j == 2:
                inc = inrows[j]
                try:
                    province = citys[inc.value]['province']
                except:
                    province = '产地(省级)'
                outc.value = province
            else:
                inc = inrows[j - 1]
                outc.value = inc.value
                outsheet.cell(row=i+1, column=j+1, value=inc.value)
    
    outbook.save(PathTable.excelout)
    pass


def main():
    add_province()
    pass


if __name__ == '__main__':
    main()
